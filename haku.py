from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import datetime
from copy import copy

def vuoro_haku(start_col, start_row, end_col, end_row, sheets):
    nimet = {}
    nimet["pvm"] = ""
    nimet["vuorot"] = []
    tyovuorot = nimet["vuorot"]
    nykyinen_nimi = ""
    
    # käydään välilehdet läpi
    for sheet_index, sheet in enumerate(sheets):
        # käydään rivit läpi
        for r in range(start_row, end_row + 1):
            # Lopetetaan jos koko rivi on tyhjä
            lopetus = True
            for u in range(start_col, end_col):
                if str(sheet.cell(r,u).value).strip() != "None":               
                    lopetus = False
                    break
            if lopetus:
                break

            # käydään sarakkeet läpi
            for c in range(start_col, end_col+1):
                # solun arvo (stripattu string)
                solu = str(sheet.cell(r,c).value).strip()
                # lisätään päivämäärä (jos 1. välilehti)
                if sheet_index == 0 and r == 3 and c == 9:
                    nimet["pvm"] = solu

                # nimirivi. käsitellään vain riviltä 6 alkaen, josta työvuorot alkavat
                if c == 1 and r > 5:                    
                    # lopetetaan rivin käsittely, jos nimirivi on tyhjä
                    if solu == "":
                        break
                    # Jos nimi löytyy jo, poistetaan sen tiedot.
                    for i, tyovuoro in enumerate(tyovuorot):
                        if list(tyovuoro.keys())[0] == solu:
                            del nimet["vuorot"][i]
                            break
                            
                    # luodaan uusi nimi
                    nimi_dict = {solu:[]}
                    tyovuorot.append(nimi_dict)
                    nykyinen_nimi = solu
                # muut rivit
                else:
                    for tyo in tyovuorot:
                        if nykyinen_nimi in tyo:
                            # useampia työaikoja päivässä: (= seuraavan rivin nimi on tyhjä, mutta sarakkeen arvo ei)
                            if str(sheet.cell(r+1, 1).value).strip() == "" and str(sheet.cell(r+1, c).value).strip() != "":
                                vuoro_dict = []
                                # ensimmäinen työaika lisätään listalle
                                vuoro_dict.append(solu)
                                # loput työajat
                                for o in range(r+1, r+20):
                                    # luupataan kunnes seuraava nimi tai työajat loppuvat
                                    if str(sheet.cell(o, 1).value).strip() == "" and str(sheet.cell(o, c).value).strip() != "":
                                        vuoro_dict.append(str(sheet.cell(o, c).value).strip())
                                    else:
                                        break
                                # lisätään työajat-lista nimet-listalle
                                tyo[nykyinen_nimi].append(vuoro_dict)
                            # jos vain yksi työaika
                            else:
                                tyo[nykyinen_nimi].append(solu)#
    return nimet

def kopioi_tyylit (default_sheet, new_sheet):
    for row in default_sheet.rows:
        for cell in row:
            new_cell = new_sheet.cell(row=cell.row, column=cell.col_idx,
                    value= cell.value)
            new_cell.font = copy(cell.font)
            new_cell.border = copy(cell.border)
            new_cell.fill = copy(cell.fill)
            new_cell.number_format = copy(cell.number_format)
            new_cell.protection = copy(cell.protection)
            new_cell.alignment = copy(cell.alignment)

def tulo_ja_meno (vuoro):
    vuoro_edit = ""
    meno = ""
    # Jos sak tai koulutus
    if vuoro[0] == "\\":
        meno = "SAK"
        vuoro_edit = vuoro[2:]
    elif vuoro[0] == "K":
        meno = "K"
        vuoro_edit = vuoro[2:]
    else:
        vuoro_edit = vuoro
    # Erotellaan tulo ja meno -ajoiksi
    if vuoro_edit[4] == "-":
        tulo = vuoro_edit[0:4]
    else:
        tulo = vuoro_edit[0:5]
    if vuoro_edit[-5] == "-":
        lahto = vuoro_edit[-4:]
    else:
        lahto = vuoro_edit[-5:]
    return tulo, lahto, meno


def laske_ajat(tyovuorot, kellonajat, valilehti, ryhma_dict, viikko_index, tyontekija, i):
    vuoro_rivi = ""
    ajat_arr = []
    for index_tyovuoro, vuoro in enumerate(tyovuorot):
        sak = False
        koulutus = False
        if kellonajat == "yhtenaiset":
            ajat = tulo_ja_meno(vuoro)
            ajat_arr.append(ajat)
            
        elif kellonajat != "erilliset":
            if index_tyovuoro  != len(tyovuorot)-1:
                vuoro_rivi += vuoro + "\n"
            else:
                vuoro_rivi += vuoro
                
    if kellonajat == "yhtenaiset":
        for ajat_index, ajat_tuple in enumerate(ajat_arr):
            # Jos ei ole viimeinen aika
            if ajat_tuple[2] != "":
                menot_str = f"{tyontekija['lempinimi']} {ajat_tuple[2]}:\n{ajat_tuple[0]}-{ajat_tuple[1]}"
                kohde = valilehti.cell(row=ryhma_dict['menot_rivi'], column=i+2-viikko_index * 7)
                if kohde.value is not None:
                    kohde.value += "\n" + menot_str
                else:
                    kohde.value = menot_str
                kohde.alignment = Alignment(wrapText=True, vertical='top', horizontal='center')
            
            if ajat_index == 0:
                alku_klo = ajat_tuple[0]
            else:
                # lähtöaika == seuraavan klo tulo-aika
                if ajat_arr[ajat_index-1][1] == ajat_tuple[0]:
                    vuoro_rivi = f"{alku_klo}-{ajat_tuple[1]}"                    
         
    return vuoro_rivi

def viikon_tyolaiset(ryhma_dict, tyovuorot_data, valilehti, viikko_index):
    # käydään läpi ryhmän työntekijöitä
    lopetus = False
    for tyontekija in ryhma_dict['tyontekijat']:
        lopetus = False
        # käydään läpi työvuoroja
        for tyyppi in tyovuorot_data:
            if lopetus:
                break
            # työvuorot löytyvät
            if tyontekija['kokonimi'] == next(iter(tyyppi)):
                for key, value in tyyppi.items():
                    for i, tyovuorot in enumerate(value):
                        # Monesko viikko menossa (0-2)
                        if i >= viikko_index * 7:
                            cell_tyovuoro = valilehti.cell(row=tyontekija['rivi'], column=i+2-viikko_index * 7)
                            if i == ryhma_dict['paivat'] + viikko_index * 7:
                                lopetus = True
                                break
                            # lempinimi
                            valilehti.cell(row=tyontekija['rivi'], column=1).value = tyontekija['lempinimi']
                            # jos useampi työaika, käydään ne läpi
                            if isinstance(tyovuorot, list):
                                vuoro_rivi = laske_ajat(tyovuorot, ryhma_dict['kellonajat'], valilehti, ryhma_dict, viikko_index, tyontekija, i)
                                cell_tyovuoro.value = vuoro_rivi                           
                            # vain yksi työaika    
                            else:
                                cell_tyovuoro.value = tyovuorot
                                
                            cell_tyovuoro.alignment = Alignment(wrapText=True, vertical='top', horizontal='center')                             
 
def tulosta_ryhma (ryhma_dict):
    data = vuoro_haku(1, 3, 22, 200, tulostettavat)
    data_pvm = data['pvm'][0:8]
    pvm = datetime.datetime.strptime(data_pvm, '%d.%m.%y')
    
    tyovuorot_data = data['vuorot']
    tiedoston_nimi = ryhma_dict['tiedosto']
    tiedosto = load_workbook(tiedoston_nimi)

    # Kirjoitetaan 3 viikkoa pedalappuja
    for viikko_index in range(0,3):
        viikko_kerroin = pvm + datetime.timedelta(days=7*viikko_index)
        viikko = viikko_kerroin.isocalendar().week
        tab_nimi = ryhma_dict['valilehti'] + '_' + str(viikko)
    
        if tab_nimi not in valilehdet:
            valilehti = tiedosto.copy_worksheet(tiedosto[ryhma_dict['nimi']])
            valilehti.title = tab_nimi
            kopioi_tyylit(tiedosto[ryhma_dict['nimi']], tiedosto[tab_nimi])
        else:
            valilehti = tiedosto[tab_nimi]
        
        viikon_tyolaiset(ryhma_dict, tyovuorot_data, valilehti, viikko_index)
    
    tiedosto.save(tiedoston_nimi)
    
def lisaa_vuorot(ryhman_nimi):
    for ryhma in ryhmat:
        if ryhma['nimi'] == ryhman_nimi:
            tulosta_ryhma(ryhma)

ryhmat = [{"nimi":"Mustikat",
           "tiedosto":"vuorot.xlsx",
           "valilehti":"Mustikat",
           "paivat":5,
           "menot_rivi":5,
           "viikkomerkinta":(1,2),
           "kellonajat":"yhtenaiset",
           "tyontekijat":[{"kokonimi":"Peppi Pitkä Paappanen",
                           "lempinimi":"Peppi",
                           "rivi":2},
                          {"kokonimi":"Bulma Baarii Leivonen",
                           "lempinimi":"Bulma",
                           "rivi":3}]}]

wb = load_workbook('vuorot.xlsx')
valilehdet = wb.sheetnames
# 2 ensimmäistä välilehteä
tulostettavat = [wb[valilehdet[0]], wb[valilehdet[1]]]

vuoro_haku(1, 3, 22, 200, tulostettavat)

# lisätään työvuorot kaikkiin ryhmiin
for ryhma in ryhmat:
    if ryhma["nimi"] != "asetukset":
        lisaa_vuorot(ryhma["nimi"])
      
